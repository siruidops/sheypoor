#!/usr/bin/env python3

import requests
from bs4 import BeautifulSoup
import time
import datetime as dt
import sys
import os
from stem import Signal
from stem.control import Controller
from openpyxl import load_workbook
from openpyxl import Workbook

url_sheypoor = "https://www.sheypoor.com"
urls = []

datenow = dt.datetime.now()
year = datenow.year; month = datenow.month; day = datenow.day
timenow = "{}_{}_{}".format(year,month,day)


def iden_():
    control_port = 9051
    with Controller.from_port(port=control_port) as controller:
        controller.authenticate()
        time.sleep(controller.get_newnym_wait())
        controller.signal(Signal.NEWNYM)


def runner():
    city = 'ایران'

    if not os.path.isfile('sheypoor-car-{}.xlsx'.format(timenow)):
        workbook_car = Workbook()
        sh_car = workbook_car.active
        sh_car['A1'] = 'ID'
        sh_car['B1'] = 'Group'
        sh_car['C1'] = 'Title'; sh_car.column_dimensions['C'].width = 30
        sh_car['D1'] = 'Url'
        sh_car['E1'] = 'Location'
        sh_car['F1'] = 'Production year'
        sh_car['G1'] = 'Model'
        sh_car['H1'] = 'kilometre'
        sh_car['I1'] = 'Color'
        sh_car['J1'] = 'Gearbox'
        sh_car['K1'] = 'Fuel'
        sh_car['L1'] = 'Body'
        sh_car['M1'] = 'Chassis'
        sh_car['N1'] = 'Description'; sh_car.column_dimensions['N'].width = 70
        sh_car['O1'] = 'Phone'; sh_car.column_dimensions['O'].width = 18
        sh_car['P1'] = 'Price'; sh_car.column_dimensions['P'].width = 15
        sh_car['Q1'] = 'Date'
        sh_car['R1'] = 'Pictures'

    else:
        workbook_car = load_workbook('sheypoor-car-{}.xlsx'.format(timenow))
        sh_car = workbook_car.worksheets[0]
        for i in range(2, sh_car.max_row+1):
            urls.append(sh_car.cell(row=i, column=4).value.strip())


    if not os.path.isfile('sheypoor-motor-{}.xlsx'.format(timenow)):
        workbook_motor = Workbook()
        sh_motor = workbook_motor.active
        sh_motor['A1'] = 'ID'
        sh_motor['B1'] = 'Group'
        sh_motor['C1'] = 'Title'; sh_motor.column_dimensions['C'].width = 30
        sh_motor['D1'] = 'Url'
        sh_motor['E1'] = 'Location'
        sh_motor['F1'] = 'Production year'
        sh_motor['G1'] = 'Engine capacity'
        sh_motor['H1'] = 'Description'; sh_motor.column_dimensions['H'].width = 70
        sh_motor['I1'] = 'Phone'; sh_motor.column_dimensions['I'].width = 18
        sh_motor['J1'] = 'Price'; sh_motor.column_dimensions['J'].width = 15
        sh_motor['K1'] = 'Date'
        sh_motor['L1'] = 'Pictures'

    else:
        workbook_motor = load_workbook('sheypoor-motor-{}.xlsx'.format(timenow))
        sh_motor = workbook_motor.worksheets[0]
        for i in range(2, sh_motor.max_row+1):
            urls.append(sh_motor.cell(row=i, column=4).value.strip())

      
    if not os.path.isfile('sheypoor-{}.xlsx'.format(timenow)):
        workbook = Workbook()
        sh = workbook.active
        sh['A1'] = 'ID'
        sh['B1'] = 'Group'
        sh['C1'] = 'Title'
        sh.column_dimensions['C'].width = 30
        sh['D1'] = 'Url'
        sh['E1'] = 'Location'
        sh['F1'] = 'Description'
        sh.column_dimensions['F'].width = 70
        sh['G1'] = 'Phone'
        sh.column_dimensions['G'].width = 30
        sh['H1'] = 'Price'
        sh.column_dimensions['G'].width = 20
        sh['I1'] = 'Date'
        sh['J1'] = 'Pictures'

    else:
        workbook = load_workbook('sheypoor-{}.xlsx'.format(timenow))
        sh = workbook.worksheets[0]
        for i in range(2, sh.max_row+1):
            urls.append(sh.cell(row=i, column=4).value.strip())

    url_city = "{}/{}". format(url_sheypoor,city)


    while 1:
        url_status_re = 0

        for i in range(1,41):
            if url_status_re == 1:
                break

            r = requests.Session()
            r.proxies = {'http':"socks5://localhost:9050",'https':"socks5://localhost:9050"}
            r.headers = {'Connection': "close", "Accept": "*/*", "Content-type": "application/x-www-form-urlencoded; charset=UTF-8", "Cookie2": "$Version=1", "Accept-Language": "en-US", "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.74 Safari/537.36 Edg/79.0.309.43"}
            success = 0

            while not success:
                try:
                    rr = r.post("{}?p={}&q=موتور".format(url_city,str(i)))
                    success = 1
                except:
                    iden_()

            source_html = rr.text
            bs = BeautifulSoup(source_html,'html.parser')
            dom = bs.find_all("div", {'class':'image'})

            for i in dom:
                lurl = i.find("a")['href']

                if lurl not in urls:
                    urls.append(lurl)
                    url_status_re = 0
                else:
                    url_status_re = 1
                    continue

                success = 0
                k = None
                while not success:
                    try:
                        k = r.post(lurl)
                        success = 1
                    except:
                        iden_()

                bs_ = BeautifulSoup(k.text,'html.parser')

                try:
                    id_ = bs_.find('meta', {'name':'listing-id'})['content']
                except:
                    continue

                success = 0
                while not success:
                    try:
                        kdes = r.get('https://www.sheypoor.com/api/web/listings/{}/description'.format(id_)).json()['data']['description']
                        success = 1
                    except:
                        iden_()

                description = kdes.split('<br />')[0].strip()
                text = """_____________________
کاربر گرامی شیپور
با توجه به دستور مقام قضایی، امکان نمایش قیمت ها وجود ندارد.
با پوزش نسبت به مشکلاتی که این اقدام برای شما به وجود آورده، تا پایان این محدودیت ناخواسته، برای اطلاع از قیمت با فروشنده تماس بگیرید."""
                description = description.split(text)[0]
                success = 0

                while not success:
                    try:
                        phone_number = r.get('https://www.sheypoor.com/api/web/listings/{}/number'.format(id_)).json()['data']['mobileNumber'].strip()
                        success = 1
                    except:
                        iden_()

                title = bs_.find("meta", {'property':'og:title'})['content'].strip()
                title = ' - '.join(title.split(' - ')[0:-1]).strip()
                image_tag = bs_.find("div", {'class':'slideshow'})
                pictures = []

                try:
                    for i in image_tag.find_all('img'):
                        try:
                            pictures.append(i['src'])
                        except KeyError:
                            pictures.append(i['data-src'])
                except:
                    pass

                pictures = ' , '.join(pictures)
                datetime = bs_.find('time').text.strip()

                try:
                    price = bs_.find('section', {'id':"item-details"})
                    price = price.find('p',{'class':"text-left"})
                    price = price.find('span',{'class':"item-price"})
                    price = price.find('strong').text.strip()
                except:
                    price = ' '

                group1 = bs_.find('meta', {'name':"prev-next.back-url"})["content"]
                group = group1.split('/')[-2]

                if '/' in group:
                    group = group.split('/')[0]

                location = bs_.find('span', {'class':"small-text"}).text.strip()

                if '/خودرو/' in group1 or '/موتور-سیکلت' in group1 :
                    production_year = ' '
                    engine_size = ' '
                    model_car = ' '
                    kilometer = ' '
                    color = ' '
                    gearbox = ' '
                    sokht = ' '
                    badane = ' '
                    shasi = ' '
                    soup = bs_.find_all('table', {"class":"key-val"})
                    j = []
                    for i in soup:
                        for jj in i.find_all('th'):
                            j.append(jj)
                            
                            
                    for i in j:
                        if i.text=='سال تولید':
                            production_year = i.find_next().text.strip()
                        elif i.text=='حجم موتور':
                            engine_size = i.find_next().text.strip()
                        elif i.text=='مدل خودرو':
                            model_car = i.find_next().text.strip()
                        elif i.text=='کیلومتر':
                            kilometer = i.find_next().text.strip()
                        elif i.text=='رنگ':
                            color = i.find_next().text.strip()
                        elif i.text=='گیربکس':
                            gearbox = i.find_next().text.strip()
                        elif i.text=='نوع سوخت':
                            sokht = i.find_next().text.strip()
                        elif i.text=='وضعیت بدنه':
                            badane = i.find_next().text.strip()
                        elif i.text=='نوع شاسی':
                            shasi = i.find_next().text.strip()
                        elif i.text=='نقدی/اقساطی':
                            price = i.find_next().text.strip()
                        else:
                            pass

                    if '/موتور-سیکلت' in group1:
                        expens = [id_, group, title, lurl, location, production_year, engine_size, description, phone_number, price, datetime, pictures]
                        sh_motor.append(expens)
                        workbook_motor.save('sheypoor-motor-{}.xlsx'.format(timenow))

                    else:
                        if color==' ':
                            print('ok')
                            o = open('a',"w")
                            o.write(str(j))
                            o.write("\n\n")
                            o.write(str(l))
                            o.write("\n\n")
                            o.write(k.text)
                            o.close()
                        expens = [id_, group, title, lurl, location, production_year, model_car, kilometer, color, gearbox, sokht, badane, shasi, description, phone_number, price, datetime, pictures]
                        sh_car.append(expens)
                        workbook_car.save('sheypoor-car-{}.xlsx'.format(timenow))

                else:
                    expens = [id_, group, title, lurl, location, description, phone_number, price, datetime, pictures]
                    sh.append(expens)
                    workbook.save('sheypoor-{}.xlsx'.format(timenow))


        time.sleep(21600)


if __name__ == "__main__":
    runner()



